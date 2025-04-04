# import openpyxl
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# import os

# def write_to_excel(rows, excel_path, sheet_name="Sheet1"):
#     if not rows:
#         return

#     if os.path.exists(excel_path):
#         wb = openpyxl.load_workbook(excel_path)
#     else:
#         wb = Workbook()
#         # Remove the default sheet if it’s still unnamed
#         if "Sheet" in wb.sheetnames:
#             std = wb["Sheet"]
#             wb.remove(std)

#     ws = wb.create_sheet(title=sheet_name)

#     for row_idx, row in enumerate(rows, 1):
#         for col_idx, cell in enumerate(row, 1):
#             ws.cell(row=row_idx, column=col_idx, value=cell)

#     wb.save(excel_path)
import openpyxl
from openpyxl.styles import Alignment

def write_to_excel(table_rows, excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Table"

    for row_idx, row in enumerate(table_rows, start=1):
        for col_idx, cell in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell)
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

    wb.save(excel_path)
